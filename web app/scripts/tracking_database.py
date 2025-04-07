#!/usr/bin/env python3
"""
Tracking Database Creator

This script creates and manages the tracking database/spreadsheet for the email form system.
It provides a structure to track email status and form returns.
"""

import os
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class TrackingDatabase:
    def __init__(self, file_path):
        """
        Initialize the tracking database with the specified file path.
        
        Args:
            file_path (str): Path to the Excel tracking spreadsheet
        """
        self.file_path = file_path
        self.required_columns = [
            'Name', 'Email', 'Date Sent', 'Email Status', 
            'Date Received', 'Form Status', 'Form Path', 'Processing Status'
        ]
        
        # Create the database if it doesn't exist
        if not os.path.exists(file_path):
            self.create_new_database()
        else:
            # Validate existing database
            try:
                df = pd.read_excel(file_path)
                missing_columns = [col for col in self.required_columns if col not in df.columns]
                if missing_columns:
                    print(f"Adding missing columns to tracking database: {missing_columns}")
                    for col in missing_columns:
                        df[col] = None
                    df.to_excel(file_path, index=False)
            except Exception as e:
                print(f"Error validating tracking database: {e}")
                self.create_new_database()
    
    def create_new_database(self):
        """
        Create a new tracking database with the required structure.
        """
        # Create a new dataframe with the required columns
        df = pd.DataFrame(columns=self.required_columns)
        
        # Save the dataframe to Excel
        df.to_excel(self.file_path, index=False)
        
        # Apply formatting
        self.apply_formatting()
        
        print(f"Created new tracking database at {self.file_path}")
    
    def apply_formatting(self):
        """
        Apply formatting to the Excel spreadsheet for better readability.
        """
        # Load the workbook
        wb = load_workbook(self.file_path)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        centered_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        # Format headers
        for col in range(1, len(self.required_columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.border = border
            
            # Adjust column width based on header text
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = max(15, len(cell.value) + 2)
        
        # Save the formatted workbook
        wb.save(self.file_path)
    
    def add_recipients(self, recipients_list):
        """
        Add new recipients to the tracking database.
        
        Args:
            recipients_list (list): List of dictionaries with recipient information
            
        Returns:
            int: Number of recipients added
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Process each recipient
        added_count = 0
        for recipient in recipients_list:
            if 'Email' in recipient and recipient['Email']:
                # Check if recipient already exists
                if not df[df['Email'] == recipient['Email']].shape[0]:
                    # Create a new record
                    new_record = {
                        'Name': recipient.get('Name', ''),
                        'Email': recipient.get('Email', ''),
                        'Email Status': 'Not Sent',
                        'Form Status': 'Not Returned',
                        'Processing Status': 'Not Started'
                    }
                    
                    # Add the record to the dataframe
                    df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)
                    added_count += 1
        
        # Save the updated database
        if added_count > 0:
            df.to_excel(self.file_path, index=False)
            self.apply_formatting()
            print(f"Added {added_count} new recipients to the tracking database")
        
        return added_count
    
    def update_email_status(self, email, status, sent_date=None):
        """
        Update the email status for a recipient.
        
        Args:
            email (str): Recipient's email address
            status (str): New email status ('Sent', 'Failed', 'Error')
            sent_date (datetime, optional): Date when the email was sent
            
        Returns:
            bool: True if the update was successful, False otherwise
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Find the recipient
        recipient_idx = df[df['Email'] == email].index
        
        if len(recipient_idx) > 0:
            # Update the status
            df.loc[recipient_idx[0], 'Email Status'] = status
            
            # Update the sent date if provided
            if sent_date and status == 'Sent':
                if isinstance(sent_date, datetime.datetime):
                    df.loc[recipient_idx[0], 'Date Sent'] = sent_date
                else:
                    df.loc[recipient_idx[0], 'Date Sent'] = datetime.datetime.now()
            
            # Save the updated database
            df.to_excel(self.file_path, index=False)
            return True
        
        return False
    
    def update_form_status(self, email, status, form_path=None, received_date=None):
        """
        Update the form status for a recipient.
        
        Args:
            email (str): Recipient's email address
            status (str): New form status ('Returned', 'Not Returned')
            form_path (str, optional): Path to the returned form
            received_date (datetime, optional): Date when the form was received
            
        Returns:
            bool: True if the update was successful, False otherwise
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Find the recipient
        recipient_idx = df[df['Email'] == email].index
        
        if len(recipient_idx) > 0:
            # Update the status
            df.loc[recipient_idx[0], 'Form Status'] = status
            
            # Update the form path if provided
            if form_path and status == 'Returned':
                df.loc[recipient_idx[0], 'Form Path'] = form_path
            
            # Update the received date if provided
            if received_date and status == 'Returned':
                if isinstance(received_date, datetime.datetime):
                    df.loc[recipient_idx[0], 'Date Received'] = received_date
                else:
                    df.loc[recipient_idx[0], 'Date Received'] = datetime.datetime.now()
            
            # Save the updated database
            df.to_excel(self.file_path, index=False)
            return True
        
        return False
    
    def update_processing_status(self, email, status):
        """
        Update the processing status for a recipient's form.
        
        Args:
            email (str): Recipient's email address
            status (str): New processing status ('Not Started', 'In Progress', 'Completed', 'Error')
            
        Returns:
            bool: True if the update was successful, False otherwise
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Find the recipient
        recipient_idx = df[df['Email'] == email].index
        
        if len(recipient_idx) > 0:
            # Update the status
            df.loc[recipient_idx[0], 'Processing Status'] = status
            
            # Save the updated database
            df.to_excel(self.file_path, index=False)
            return True
        
        return False
    
    def get_recipients_by_status(self, email_status=None, form_status=None, processing_status=None):
        """
        Get recipients filtered by various status fields.
        
        Args:
            email_status (str, optional): Filter by email status
            form_status (str, optional): Filter by form status
            processing_status (str, optional): Filter by processing status
            
        Returns:
            pd.DataFrame: Filtered dataframe of recipients
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Apply filters
        if email_status:
            df = df[df['Email Status'] == email_status]
        
        if form_status:
            df = df[df['Form Status'] == form_status]
        
        if processing_status:
            df = df[df['Processing Status'] == processing_status]
        
        return df
    
    def generate_status_report(self, output_file=None):
        """
        Generate a status report of the tracking database.
        
        Args:
            output_file (str, optional): Path to save the report
            
        Returns:
            dict: Status report statistics
        """
        # Load the current database
        df = pd.read_excel(self.file_path)
        
        # Calculate statistics
        total_recipients = len(df)
        emails_sent = len(df[df['Email Status'] == 'Sent'])
        emails_failed = len(df[df['Email Status'] == 'Failed'])
        forms_returned = len(df[df['Form Status'] == 'Returned'])
        forms_processed = len(df[df['Processing Status'] == 'Completed'])
        
        # Create report
        report = {
            'Total Recipients': total_recipients,
            'Emails Sent': emails_sent,
            'Emails Failed': emails_failed,
            'Forms Returned': forms_returned,
            'Forms Processed': forms_processed,
            'Sent Percentage': f"{(emails_sent / total_recipients * 100) if total_recipients > 0 else 0:.1f}%",
            'Return Rate': f"{(forms_returned / emails_sent * 100) if emails_sent > 0 else 0:.1f}%",
            'Processing Rate': f"{(forms_processed / forms_returned * 100) if forms_returned > 0 else 0:.1f}%"
        }
        
        # Save report to file if requested
        if output_file:
            with open(output_file, 'w') as f:
                f.write("# Email Form Tracking Status Report\n\n")
                f.write(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                for key, value in report.items():
                    f.write(f"- **{key}**: {value}\n")
        
        return report

if __name__ == "__main__":
    print("Tracking Database Creator - Use this module by importing it in your main script")
    print("Example usage:")
    print("  from tracking_database import TrackingDatabase")
    print("  db = TrackingDatabase('tracking.xlsx')")
    print("  db.add_recipients([{'Name': 'John Doe', 'Email': 'john@example.com'}])")
    print("  db.update_email_status('john@example.com', 'Sent')")
