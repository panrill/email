#!/usr/bin/env python3
"""
Sample Tracking Spreadsheet Creator

This script creates a sample tracking spreadsheet with example recipients
for demonstration purposes.
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_sample_tracking_spreadsheet(output_file):
    """
    Create a sample tracking spreadsheet with example recipients.
    
    Args:
        output_file (str): Path to save the tracking spreadsheet
        
    Returns:
        str: Path to the created spreadsheet
    """
    # Create a dataframe with sample data
    sample_data = [
        {
            'Name': 'John Smith',
            'Email': 'john.smith@example.com',
            'Date Sent': None,
            'Email Status': 'Not Sent',
            'Date Received': None,
            'Form Status': 'Not Returned',
            'Form Path': None,
            'Processing Status': 'Not Started'
        },
        {
            'Name': 'Jane Doe',
            'Email': 'jane.doe@example.com',
            'Date Sent': None,
            'Email Status': 'Not Sent',
            'Date Received': None,
            'Form Status': 'Not Returned',
            'Form Path': None,
            'Processing Status': 'Not Started'
        },
        {
            'Name': 'Robert Johnson',
            'Email': 'robert.johnson@example.com',
            'Date Sent': None,
            'Email Status': 'Not Sent',
            'Date Received': None,
            'Form Status': 'Not Returned',
            'Form Path': None,
            'Processing Status': 'Not Started'
        }
    ]
    
    df = pd.DataFrame(sample_data)
    
    # Save the dataframe to Excel
    df.to_excel(output_file, index=False)
    
    # Apply formatting
    apply_formatting(output_file)
    
    print(f"Created sample tracking spreadsheet at {output_file}")
    return output_file

def apply_formatting(file_path):
    """
    Apply formatting to the Excel spreadsheet for better readability.
    
    Args:
        file_path (str): Path to the Excel spreadsheet
    """
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    centered_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), 
        right=Side(style="thin"), 
        top=Side(style="thin"), 
        bottom=Side(style="thin")
    )
    
    # Format headers
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border
        
        # Adjust column width based on header text
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = max(15, len(cell.value) + 2)
    
    # Format data cells
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            
            # Center status columns
            if col in [4, 6, 8]:  # Email Status, Form Status, Processing Status
                cell.alignment = centered_alignment
    
    # Save the formatted workbook
    wb.save(file_path)

if __name__ == "__main__":
    # Create a sample tracking spreadsheet in the data directory
    data_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data')
    os.makedirs(data_dir, exist_ok=True)
    
    output_file = os.path.join(data_dir, 'sample_tracking.xlsx')
    create_sample_tracking_spreadsheet(output_file)
    
    print(f"Sample tracking spreadsheet created at: {output_file}")
