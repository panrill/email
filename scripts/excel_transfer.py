#!/usr/bin/env python3
"""
Excel Data Transfer Script

This script transfers data extracted from PDF forms to Excel spreadsheets.
It can process individual JSON data files or batch process multiple files,
and supports various Excel formatting options.
"""

import os
import sys
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class ExcelDataTransfer:
    def __init__(self, output_file):
        """
        Initialize the Excel data transfer with the output file path.
        
        Args:
            output_file (str): Path to the Excel file to create or update
        """
        self.output_file = output_file
        self.data = []
        self.columns = []
    
    def add_data_from_json(self, json_file):
        """
        Add data from a JSON file to the dataset.
        
        Args:
            json_file (str): Path to the JSON file with extracted form data
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Load JSON data
            with open(json_file, 'r') as f:
                data = json.load(f)
            
            # Extract relevant data
            record = {}
            
            # Add form fields
            if 'form_fields' in data and data['form_fields']:
                record.update(data['form_fields'])
            
            # Add extracted data
            if 'extracted_data' in data and data['extracted_data']:
                record.update(data['extracted_data'])
            
            # Add metadata
            if 'metadata' in data:
                record['filename'] = data['metadata'].get('filename', '')
                record['extraction_methods'] = ','.join(data['metadata'].get('extraction_methods', []))
            
            # Add table data as separate records if present
            if 'table_data' in data and data['table_data']:
                for i, table_row in enumerate(data['table_data']):
                    table_record = record.copy()
                    table_record.update({f"table_{k}": v for k, v in table_row.items()})
                    table_record['record_type'] = 'table_row'
                    table_record['row_number'] = i + 1
                    self.data.append(table_record)
                    
                    # Update columns
                    self.columns = list(set(self.columns) | set(table_record.keys()))
            else:
                # Add as single record
                record['record_type'] = 'form'
                self.data.append(record)
                
                # Update columns
                self.columns = list(set(self.columns) | set(record.keys()))
            
            return True
        
        except Exception as e:
            print(f"Error processing JSON file {json_file}: {e}")
            return False
    
    def add_data_from_directory(self, json_dir):
        """
        Add data from all JSON files in a directory.
        
        Args:
            json_dir (str): Directory containing JSON files with extracted form data
            
        Returns:
            int: Number of files successfully processed
        """
        # Get all JSON files in the directory
        json_files = [f for f in os.listdir(json_dir) if f.lower().endswith('.json')]
        
        if not json_files:
            print(f"No JSON files found in {json_dir}")
            return 0
        
        # Process each file
        success_count = 0
        for json_file in json_files:
            json_path = os.path.join(json_dir, json_file)
            if self.add_data_from_json(json_path):
                success_count += 1
        
        return success_count
    
    def create_excel(self, include_timestamp=True):
        """
        Create an Excel file with the collected data.
        
        Args:
            include_timestamp (bool): Whether to include a timestamp in the filename
            
        Returns:
            str: Path to the created Excel file
        """
        if not self.data:
            print("No data to export")
            return None
        
        # Create DataFrame
        df = pd.DataFrame(self.data)
        
        # Reorder columns for better readability
        priority_columns = ['record_type', 'row_number', 'name', 'email', 'phone', 'date', 'address', 'filename', 'extraction_methods']
        available_priority_columns = [col for col in priority_columns if col in df.columns]
        other_columns = [col for col in df.columns if col not in priority_columns]
        df = df[available_priority_columns + other_columns]
        
        # Determine output file path
        output_file = self.output_file
        if include_timestamp:
            base, ext = os.path.splitext(output_file)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = f"{base}_{timestamp}{ext}"
        
        # Save to Excel
        df.to_excel(output_file, index=False)
        
        # Apply formatting
        self.apply_formatting(output_file)
        
        return output_file
    
    def apply_formatting(self, excel_file):
        """
        Apply formatting to the Excel file for better readability.
        
        Args:
            excel_file (str): Path to the Excel file to format
        """
        # Load workbook
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        centered_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"), 
            right=Side(style="thin"), 
            top=Side(style="thin"), 
            bottom=Side(style="thin")
        )
        
        # Format headers
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.border = border
            
            # Adjust column width based on header text
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = max(15, len(str(cell.value)) + 2)
        
        # Format data cells
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                # Center specific columns
                if col == 1:  # record_type
                    cell.alignment = centered_alignment
        
        # Save the formatted workbook
        wb.save(excel_file)
    
    def update_tracking_spreadsheet(self, tracking_file, email_column='Email', status_column='Processing Status'):
        """
        Update the tracking spreadsheet with processing status.
        
        Args:
            tracking_file (str): Path to the tracking spreadsheet
            email_column (str): Column name for email addresses
            status_column (str): Column name for processing status
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Load tracking spreadsheet
            tracking_df = pd.read_excel(tracking_file)
            
            # Check if required columns exist
            if email_column not in tracking_df.columns or status_column not in tracking_df.columns:
                print(f"Required columns not found in tracking spreadsheet: {email_column}, {status_column}")
                return False
            
            # Get emails from processed data
            processed_emails = set()
            for record in self.data:
                if 'email' in record and record['email']:
                    processed_emails.add(record['email'])
            
            # Update status for processed emails
            for email in processed_emails:
                idx = tracking_df[tracking_df[email_column] == email].index
                if len(idx) > 0:
                    tracking_df.loc[idx[0], status_column] = 'Completed'
            
            # Save updated tracking spreadsheet
            tracking_df.to_excel(tracking_file, index=False)
            
            return True
        
        except Exception as e:
            print(f"Error updating tracking spreadsheet: {e}")
            return False

def process_extracted_data(input_path, output_file, tracking_file=None):
    """
    Process extracted data from JSON files and transfer to Excel.
    
    Args:
        input_path (str): Path to JSON file or directory with JSON files
        output_file (str): Path to save the Excel file
        tracking_file (str, optional): Path to tracking spreadsheet to update
        
    Returns:
        str: Path to the created Excel file
    """
    # Initialize Excel data transfer
    transfer = ExcelDataTransfer(output_file)
    
    # Process input
    if os.path.isdir(input_path):
        # Process directory
        file_count = transfer.add_data_from_directory(input_path)
        print(f"Processed {file_count} JSON files from {input_path}")
    else:
        # Process single file
        if transfer.add_data_from_json(input_path):
            print(f"Processed JSON file: {input_path}")
        else:
            print(f"Failed to process JSON file: {input_path}")
            return None
    
    # Create Excel file
    excel_file = transfer.create_excel()
    if excel_file:
        print(f"Data transferred to Excel file: {excel_file}")
        
        # Update tracking spreadsheet if provided
        if tracking_file and os.path.exists(tracking_file):
            if transfer.update_tracking_spreadsheet(tracking_file):
                print(f"Updated tracking spreadsheet: {tracking_file}")
            else:
                print(f"Failed to update tracking spreadsheet: {tracking_file}")
    
    return excel_file

if __name__ == "__main__":
    # Parse command line arguments
    if len(sys.argv) < 3:
        print("Usage: python excel_transfer.py <json_file_or_dir> <output_excel> [tracking_file]")
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_file = sys.argv[2]
    tracking_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    # Process data
    excel_file = process_extracted_data(input_path, output_file, tracking_file)
    
    if not excel_file:
        print("Data transfer failed")
        sys.exit(1)
